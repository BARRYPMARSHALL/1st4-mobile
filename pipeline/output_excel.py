"""
1st 4 Mobile — Dispute Schedule Excel Generation

Generates a professionally formatted Excel dispute schedule with
6 sheets: Executive Summary, Ghost Lines, Rate Mismatches, Roaming
Anomalies, Legacy Rollbacks, and Duplicates.

Uses openpyxl for all formatting (not xlsxwriter).
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side
from openpyxl.utils import get_column_letter

from pipeline.config import DEFAULT_HISTORICAL_MONTHS, OUTPUT_DATE_FORMAT
from pipeline.utils.money_utils import format_currency
from pipeline.utils.logging_utils import AuditLogger

logger = logging.getLogger("1st4pipeline.output_excel")

# ── Style constants ───────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
BORDER_THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CURRENCY_FMT = '$#,##0.00'
PCT_FMT = '0.0%'
DATE_FMT = 'YYYY-MM-DD'

# Columns that should be formatted as currency in each sheet
CURRENCY_COLUMNS = {
    "Executive Summary": {"monthly_overcharge", "annualised_overcharge"},
    "Ghost Lines": {"monthly_overcharge", "12mo_total"},
    "Rate Mismatches": {"contracted_rate", "billed_rate", "variance", "variance_pct"},
    "Roaming Anomalies": {"contracted_rate", "billed_rate"},
    "Legacy Rollbacks": {"previous_rate", "current_rate", "monthly_variance"},
    "Duplicates": {"charge_amount"},
}

PCT_COLUMNS = {
    "Rate Mismatches": {"variance_pct"},
}


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Auto-fit column widths based on cell content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            # Estimate character width (handling wider chars roughly)
            cell_len = len(val)
            if any(ord(c) > 127 for c in val):
                cell_len *= 1.3  # wider CJK/unicode chars
            if cell_len > max_len:
                max_len = cell_len
        adjusted = max(min_width, min(max_len + 3, max_width))
        ws.column_dimensions[col_letter].width = adjusted


def _style_header_row(ws) -> None:
    """Apply bold white-on-blue header styling."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN


def _apply_alt_row_colors(ws, start_row: int = 2) -> None:
    """Apply alternating row background colours."""
    for row_idx in range(start_row, ws.max_row + 1):
        if (row_idx - start_row) % 2 == 1:
            for cell in ws[row_idx]:
                cell.fill = ALT_ROW_FILL


def _apply_borders(ws) -> None:
    """Apply thin borders to all data cells."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = BORDER_THIN


def _apply_currency_format(ws, sheet_name: str, header_row: list[str]) -> None:
    """Apply currency number format to columns that contain monetary values."""
    currency_cols = CURRENCY_COLUMNS.get(sheet_name, set())
    pct_cols = PCT_COLUMNS.get(sheet_name, set())

    for col_idx, header in enumerate(header_row, start=1):
        # Normalise: lowercase, strip, replace spaces/special chars with _
        normalized = (
            str(header).lower().strip()
            .replace(" ", "_")
            .replace("-", "_")
            .replace("(", "")
            .replace(")", "")
        )
        if normalized in currency_cols:
            for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = CURRENCY_FMT
        elif normalized in pct_cols:
            for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = PCT_FMT


def _safe_value(val, default: str = "") -> str:
    """Return a string representation, handling None/NaN."""
    if val is None:
        return default
    if isinstance(val, float) and pd.isna(val):
        return default
    return str(val)


def _build_summary_rows(all_flags: dict):
    """Build the Executive Summary rows from the all_flags dict.

    Yields [category, count, monthly_overcharge, annualised_overcharge, avg_confidence].
    """
    summary = all_flags.get("summary", {})
    breakdown = summary.get("breakdown", {})
    monthly_breakdown = summary.get("monthly_breakdown", {})
    total_monthly = float(summary.get("total_monthly_overcharge", 0))
    total_annualised = float(summary.get("total_annualised", 0))
    total_flags = int(summary.get("total_flags", 0))

    # Per-engine averages / totals — we compute avg confidence from each engine's DataFrame
    engine_confidences = {}
    engine_labels = {
        "ghost_lines": "Ghost Lines",
        "rate_mismatches": "Rate Mismatches",
        "roaming": "Roaming Anomalies",
        "legacy_rollbacks": "Legacy Rollbacks",
        "duplicates": "Duplicates",
    }

    for engine_key, label in engine_labels.items():
        df = all_flags.get(engine_key, pd.DataFrame())
        count = breakdown.get(engine_key, 0)
        monthly = monthly_breakdown.get(engine_key, 0.0)
        annualised = monthly * 12.0

        avg_conf = 0.0
        if count > 0 and not df.empty and "confidence" in df.columns:
            conf_vals = df["confidence"].dropna()
            if len(conf_vals) > 0:
                avg_conf = float(conf_vals.mean())

        yield [
            label,
            count,
            round(monthly, 2),
            round(annualised, 2),
            round(avg_conf, 2),
        ]

    # Totals row
    # Overall avg confidence across all engines
    all_confidences = []
    for engine_key in engine_labels:
        df = all_flags.get(engine_key, pd.DataFrame())
        if not df.empty and "confidence" in df.columns:
            all_confidences.extend(df["confidence"].dropna().tolist())
    overall_avg_conf = round(sum(all_confidences) / len(all_confidences), 2) if all_confidences else 0.0

    yield [
        "TOTAL",
        total_flags,
        round(total_monthly, 2),
        round(total_annualised, 2),
        overall_avg_conf,
    ]


def _write_summary_sheet(wb: Workbook, all_flags: dict) -> None:
    """Write Sheet 1: Executive Summary."""
    ws = wb.create_sheet("Executive Summary", 0)
    headers = ["Category", "Count", "Monthly Overcharge", "Annualised Overcharge", "Avg Confidence"]
    ws.append(headers)

    for row_data in _build_summary_rows(all_flags):
        ws.append(row_data)

    _style_header_row(ws)
    _apply_alt_row_colors(ws)
    _apply_borders(ws)
    _apply_currency_format(ws, "Executive Summary", headers)
    _auto_width(ws)


def _write_ghost_lines_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Write Sheet 2: Ghost Lines."""
    ws = wb.create_sheet("Ghost Lines")
    headers = ["Service ID", "Detection Method", "Confidence", "Monthly Overcharge", "12mo Total", "Detail"]
    ws.append(headers)

    historical_months = DEFAULT_HISTORICAL_MONTHS

    if not df.empty:
        for _, row in df.iterrows():
            service_id = _safe_value(row.get("service_id"))
            method = _safe_value(row.get("detection_method", ""))
            confidence = float(row.get("confidence", 0) or 0)
            monthly = float(row.get("estimated_monthly_overcharge", 0) or 0)
            twelve_mo = monthly * historical_months
            detail = _safe_value(row.get("detail", ""))

            ws.append([service_id, method, confidence, monthly, twelve_mo, detail])

    _style_header_row(ws)
    _apply_alt_row_colors(ws)
    _apply_borders(ws)
    _apply_currency_format(ws, "Ghost Lines", headers)
    _auto_width(ws)


def _write_rate_mismatches_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Write Sheet 3: Rate Mismatches."""
    ws = wb.create_sheet("Rate Mismatches")
    headers = ["Plan Code", "Contracted Rate", "Billed Rate", "Variance", "Variance %"]
    ws.append(headers)

    if not df.empty:
        for _, row in df.iterrows():
            plan_code = _safe_value(row.get("plan_code", ""))
            contracted = float(row.get("contracted_amount", 0) or 0)
            billed_rate = float(row.get("billed_amount", 0) or 0)
            # variance can be from variance_amount or computed
            variance = float(row.get("variance_amount", 0) or 0)
            variance_pct = float(row.get("variance_pct", 0) or 0) / 100.0  # convert to decimal for %

            ws.append([plan_code, contracted, billed_rate, variance, variance_pct])

    _style_header_row(ws)
    _apply_alt_row_colors(ws)
    _apply_borders(ws)
    _apply_currency_format(ws, "Rate Mismatches", headers)
    _auto_width(ws)


def _write_roaming_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Write Sheet 4: Roaming Anomalies."""
    ws = wb.create_sheet("Roaming Anomalies")
    headers = ["Zone", "Charge Type", "Contracted Rate", "Billed Rate", "Usage"]
    ws.append(headers)

    if not df.empty:
        for _, row in df.iterrows():
            zone = _safe_value(row.get("zone", ""))
            charge_type = _safe_value(row.get("charge_type", "data"))
            contracted = float(row.get("contracted_rate", 0) or 0)
            billed = float(row.get("billed_rate", 0) or 0)
            usage = float(row.get("usage_units", 0) or 0)

            ws.append([zone, charge_type, contracted, billed, usage])

    _style_header_row(ws)
    _apply_alt_row_colors(ws)
    _apply_borders(ws)
    _apply_currency_format(ws, "Roaming Anomalies", headers)
    _auto_width(ws)


def _write_legacy_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Write Sheet 5: Legacy Rollbacks."""
    ws = wb.create_sheet("Legacy Rollbacks")
    headers = ["Previous Plan", "New Plan", "Previous Rate", "Current Rate", "Monthly Variance"]
    ws.append(headers)

    if not df.empty:
        for _, row in df.iterrows():
            prev_plan = _safe_value(row.get("previous_plan", ""))
            new_plan = _safe_value(row.get("current_plan", ""))
            prev_rate = float(row.get("previous_amount", 0) or 0)
            curr_rate = float(row.get("current_amount", 0) or 0)
            monthly_var = float(row.get("estimated_monthly_overcharge", 0) or 0)

            ws.append([prev_plan, new_plan, prev_rate, curr_rate, monthly_var])

    _style_header_row(ws)
    _apply_alt_row_colors(ws)
    _apply_borders(ws)
    _apply_currency_format(ws, "Legacy Rollbacks", headers)
    _auto_width(ws)


def _write_duplicates_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Write Sheet 6: Duplicates."""
    ws = wb.create_sheet("Duplicates")
    headers = ["Service ID", "Charge Description", "Charge Amount", "Charge Period"]
    ws.append(headers)

    if not df.empty:
        for _, row in df.iterrows():
            service_id = _safe_value(row.get("service_id", ""))
            desc = _safe_value(row.get("charge_description", ""))
            amount = float(row.get("charge_amount", 0) or 0)
            period = _safe_value(row.get("charge_period_start", ""))

            ws.append([service_id, desc, amount, period])

    _style_header_row(ws)
    _apply_alt_row_colors(ws)
    _apply_borders(ws)
    _apply_currency_format(ws, "Duplicates", headers)
    _auto_width(ws)


def generate_dispute_schedule(
    all_flags: dict,
    df_raw: pd.DataFrame,
    client_name: str,
    output_path: str,
    contract: object = None,
    audit: AuditLogger = None,
) -> str:
    """Generate a professionally formatted Excel dispute schedule.

    Creates a 6-sheet workbook with the Executive Summary, Ghost Lines,
    Rate Mismatches, Roaming Anomalies, Legacy Rollbacks, and Duplicates.

    Args:
        all_flags: Result dict from run_all_detections() containing
            DataFrames for each engine and a summary dict.
        df_raw: The raw (or normalised) billing DataFrame — used for
            context but not written directly to sheets.
        client_name: Client name for workbook metadata.
        output_path: Full path to write the .xlsx file.
        contract: Optional ContractMatrix instance (for metadata).
        audit: Optional AuditLogger for audit trail.

    Returns:
        The absolute path to the generated Excel file as a string.

    Raises:
        IOError: If the file cannot be written.
    """
    output_path = str(output_path)
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    logger.info(
        f"Generating dispute schedule for '{client_name}' → {output_file.name}"
    )

    wb = Workbook()

    # Remove the default sheet if it exists
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    # ── Sheet 1: Executive Summary ──────────────────────────────────
    _write_summary_sheet(wb, all_flags)

    # ── Sheet 2: Ghost Lines ────────────────────────────────────────
    ghost_df = all_flags.get("ghost_lines", pd.DataFrame())
    _write_ghost_lines_sheet(wb, ghost_df)

    # ── Sheet 3: Rate Mismatches ────────────────────────────────────
    rate_df = all_flags.get("rate_mismatches", pd.DataFrame())
    _write_rate_mismatches_sheet(wb, rate_df)

    # ── Sheet 4: Roaming Anomalies ──────────────────────────────────
    roam_df = all_flags.get("roaming", pd.DataFrame())
    _write_roaming_sheet(wb, roam_df)

    # ── Sheet 5: Legacy Rollbacks ───────────────────────────────────
    legacy_df = all_flags.get("legacy_rollbacks", pd.DataFrame())
    _write_legacy_sheet(wb, legacy_df)

    # ── Sheet 6: Duplicates ─────────────────────────────────────────
    dupe_df = all_flags.get("duplicates", pd.DataFrame())
    _write_duplicates_sheet(wb, dupe_df)

    # ── Workbook metadata ───────────────────────────────────────────
    wb.properties.creator = "1st 4 Mobile Audit Pipeline"
    wb.properties.title = f"Dispute Schedule — {client_name}"
    wb.properties.description = (
        f"Automated billing dispute schedule for {client_name}. "
        f"Generated on {datetime.now().strftime(OUTPUT_DATE_FORMAT)}."
    )

    # ── Save ────────────────────────────────────────────────────────
    try:
        wb.save(str(output_file))
    except Exception as exc:
        logger.error(f"Failed to write Excel file at {output_path}: {exc}")
        raise IOError(f"Could not write Excel file: {exc}") from exc

    logger.info(
        f"Dispute schedule saved: {output_file.resolve()} "
        f"({len(wb.sheetnames)} sheets)"
    )

    if audit:
        audit.log(
            "output_excel", "dispute_schedule",
            f"Generated dispute schedule with {len(wb.sheetnames)} sheets",
            amount=output_file.stat().st_size / 1024.0,
        )

    return str(output_file.resolve())
