"""
1st 4 Mobile — XLSX Ingestor
Excel parser for Telstra T-Analyst .xlsx exports.

Handles:
- Finding the data sheet (skipping metadata/summary sheets)
- Merged cells (expand to fill)
- Multi-header detection (same heuristic as CSV ingestor)
- Returns (DataFrame, audit_info_dict)
"""

import logging
from pathlib import Path
from typing import Optional

import pandas as pd

from pipeline.config import TELSTRA_FINGERPRINTS, OPTUS_FINGERPRINTS

logger = logging.getLogger("1st4pipeline.xlsx_ingestor")

# Sheet name patterns that indicate summary/metadata sheets to skip
SKIP_SHEET_PATTERNS = [
    "summary", "metadata", "notes", "glossary", "readme",
    "cover", "contents", "index", "legend", "toc",
]


def _is_data_sheet(sheet_name: str, df: pd.DataFrame) -> bool:
    """Determine if an Excel sheet is likely a data sheet.

    Heuristic: skip sheets with few rows/columns or those whose names
    match known skip patterns.
    """
    name_lower = sheet_name.lower().strip()
    if any(pattern in name_lower for pattern in SKIP_SHEET_PATTERNS):
        return False
    if df.shape[0] < 3 or df.shape[1] < 3:
        return False
    return True


def _find_data_sheet(file_path: Path) -> tuple[Optional[str], pd.DataFrame]:
    """Find the primary data sheet in a multi-sheet workbook.

    Returns (sheet_name, raw_dataframe) or (None, empty DataFrame).
    """
    try:
        xls = pd.ExcelFile(file_path, engine="openpyxl")
    except Exception as exc:
        logger.error(f"Failed to open {file_path.name}: {exc}")
        return None, pd.DataFrame()

    # Prefer sheets with Telstra-like data
    data_sheets = []

    for sheet_name in xls.sheet_names:
        try:
            # Read small sample to evaluate
            df_sample = pd.read_excel(
                xls, sheet_name=sheet_name, nrows=5, header=None, dtype=str
            )
            if not _is_data_sheet(sheet_name, df_sample):
                continue

            header_text = " ".join(
                str(c).lower() for c in df_sample.iloc[0] if pd.notna(c)
            )

            # Score based on Telstra/Optus fingerprints
            score = 0
            for pattern in TELSTRA_FINGERPRINTS:
                if pattern.lower() in header_text:
                    score += 2
            for pattern in OPTUS_FINGERPRINTS:
                if pattern.lower() in header_text:
                    score += 1

            row_count = df_sample.shape[0]
            data_sheets.append((sheet_name, score, row_count))
        except Exception as exc:
            logger.debug(f"Skipping sheet '{sheet_name}': {exc}")
            continue

    if not data_sheets:
        # Fall back: find the sheet with most rows
        sheets_with_rows = []
        for sheet_name in xls.sheet_names:
            try:
                df_sample = pd.read_excel(
                    xls, sheet_name=sheet_name, nrows=5, header=None, dtype=str
                )
                sheets_with_rows.append(
                    (sheet_name, df_sample.shape[0], df_sample.shape[1])
                )
            except Exception:
                continue
        if sheets_with_rows:
            sheets_with_rows.sort(key=lambda x: x[1], reverse=True)
            chosen = sheets_with_rows[0][0]
            logger.info(f"Fallback: using sheet '{chosen}' (most rows)")
        else:
            # Try first sheet
            chosen = xls.sheet_names[0] if xls.sheet_names else None
            if chosen is None:
                return None, pd.DataFrame()
            logger.info(f"Fallback: using first sheet '{chosen}'")
    else:
        # Sort by score desc, then by row count desc
        data_sheets.sort(key=lambda x: (x[1], x[2]), reverse=True)
        chosen = data_sheets[0][0]

    # Read the chosen sheet fully
    try:
        df = pd.read_excel(xls, sheet_name=chosen, header=None, dtype=str)
    except Exception as exc:
        logger.error(f"Failed to read sheet '{chosen}': {exc}")
        return chosen, pd.DataFrame()

    return chosen, df


def _expand_merged_cells(df: pd.DataFrame, file_path: Path) -> pd.DataFrame:
    """Fill merged-cell NaN values using openpyxl's merged_cells info.

    openpyxl tracks which cells are merged; we forward-fill from the
    top-left cell of each merged range across the entire range.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not available — cannot expand merged cells")
        return df

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        logger.warning(f"openpyxl load failed for merged cell detection: {exc}")
        return df

    try:
        # Find the active/data sheet that matches our DataFrame
        # We'll iterate all sheets and look for one with merged cells
        merged_ranges = []
        for ws in wb.worksheets:
            if ws.merged_cells.ranges:
                merged_ranges = list(ws.merged_cells.ranges)
                break
    finally:
        wb.close()

    if not merged_ranges:
        return df

    # Expand merged cells by forward-filling
    for merged_range in merged_ranges:
        try:
            min_col = merged_range.min_col - 1  # 0-indexed
            min_row = merged_range.min_row - 1
            max_col = merged_range.max_col - 1
            max_row = merged_range.max_row - 1

            # Get the top-left value
            if min_row < len(df) and min_col < len(df.columns):
                fill_value = df.iloc[min_row, min_col]

                # Fill the entire merged range
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        if r < len(df) and c < len(df.columns):
                            if pd.isna(df.iloc[r, c]) or df.iloc[r, c] == "":
                                df.iloc[r, c] = fill_value
        except Exception:
            continue

    return df


def _count_header_rows(df_raw: pd.DataFrame) -> int:
    """Detect whether the raw dataframe has 1 or 2 header rows.

    Uses the same heuristic as the CSV ingestor.
    """
    if df_raw.shape[0] < 2:
        return 1

    cols_line1 = [str(c).strip() if pd.notna(c) else "" for c in df_raw.iloc[0]]
    cols_line2 = [str(c).strip() if pd.notna(c) else "" for c in df_raw.iloc[1]]

    if len(cols_line1) < 3:
        return 2 if any(c for c in cols_line1) else 1

    unit_keywords = {
        "mb", "gb", "$", "inc", "gst", "excl", "units",
        "minutes", "sms", "count", "qty",
    }
    header_keywords = {
        "account", "service", "charge", "plan", "invoice",
        "description", "amount", "usage", "rate", "date",
        "period", "number", "name", "code",
    }

    line2_words = set(
        w for cell in cols_line2
        for w in cell.lower().replace("/", " ").split()
    )
    line1_words = set(
        w for cell in cols_line1
        for w in cell.lower().replace("/", " ").split()
    )

    unit_ratio = len(line2_words & unit_keywords) / max(len(line2_words), 1)
    header_ratio = len(line1_words & header_keywords) / max(len(line1_words), 1)

    return 2 if header_ratio > 0.3 and unit_ratio > 0.3 else 1


def _fingerprint_provider(headers: list[str]) -> str:
    """Detect telecom provider from column headers."""
    header_text = " ".join(h.lower() for h in headers if h)

    telstra_hits = sum(
        1 for f in TELSTRA_FINGERPRINTS if f.lower() in header_text
    )
    optus_hits = sum(
        1 for f in OPTUS_FINGERPRINTS if f.lower() in header_text
    )

    if telstra_hits > optus_hits:
        return "telstra"
    elif optus_hits > telstra_hits:
        return "optus"

    # Per-column substring matching
    for h in headers:
        hl = h.lower().strip()
        for f in TELSTRA_FINGERPRINTS:
            if f.lower() in hl:
                return "telstra"
        for f in OPTUS_FINGERPRINTS:
            if f.lower() in hl:
                return "optus"

    return "unknown"


def ingest_xlsx(file_path: str | Path) -> tuple:
    """Ingest an Excel (.xlsx) file from Telstra T-Analyst exports.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        Tuple of (DataFrame, audit_info_dict). audit_info has:
        source_file, rows, columns, provider, n_header_rows,
        sheet_name, column_names.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If no data sheet can be identified or file is empty.
    """
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {file_path}")

    # Step 1: Find the data sheet
    sheet_name, df_raw = _find_data_sheet(file_path)
    if sheet_name is None or df_raw.empty:
        raise ValueError(
            f"No data sheet found in {file_path.name}"
        )
    logger.info(f"Identified data sheet: '{sheet_name}'")

    # Step 2: Expand merged cells
    df_raw = _expand_merged_cells(df_raw, file_path)

    # Step 3: Detect header rows
    n_header_rows = _count_header_rows(df_raw)
    logger.debug(f"Detected header rows: {n_header_rows}")

    # Step 4: Extract header and data rows
    header_row = n_header_rows - 1
    raw_headers = [
        str(c).strip() if pd.notna(c) else f"col_{i}"
        for i, c in enumerate(df_raw.iloc[header_row])
    ]

    # Fingerprint provider
    provider = _fingerprint_provider(raw_headers)

    # Extract data rows
    data_start = n_header_rows
    df_data = df_raw.iloc[data_start:].reset_index(drop=True)
    df_data.columns = raw_headers

    # If we had 2 header rows, the first header row becomes an additional
    # composite header that we can optionally combine. For now, just use
    # the second header row as column names.

    # Remove entirely empty rows
    df_data = df_data.dropna(how="all").reset_index(drop=True)

    # Remove entirely empty columns
    df_data = df_data.dropna(axis=1, how="all")

    # Strip whitespace from string cells
    for col in df_data.select_dtypes(include="object").columns:
        df_data[col] = df_data[col].astype(str).str.strip()

    rows, cols = df_data.shape

    if rows == 0:
        raise ValueError(f"No data rows found in sheet '{sheet_name}'")

    audit_info = {
        "source_file": str(file_path.resolve()),
        "rows": rows,
        "columns": cols,
        "provider": provider,
        "n_header_rows": n_header_rows,
        "sheet_name": sheet_name,
        "column_names": list(df_data.columns),
    }

    logger.info(
        f"Ingested {file_path.name} sheet '{sheet_name}': "
        f"{rows} rows × {cols} cols, provider={provider}"
    )

    return df_data, audit_info
